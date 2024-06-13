# -*- coding: utf-8 -*-

import os
import pandas as pd
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
now = dt.now()
last_month = now - relativedelta(months=1)
date = last_month.strftime('%Y-%m')

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, ProjectedPieChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Color, Alignment, Border, Side, PatternFill
import subprocess

def metal_contents(table_name, df, planes):

    if planes == 'МЕРИДИОНАЛЬНАЯ_ПК':
        key = table_name.split('.')[0]
    elif planes == 'МЕРИДИОНАЛЬНАЯ_ГР_Л':
        key = table_name.split('гр.л. ')[1]
        key = key.split('.')[0]
    print(key)
    
        
    ni = (df['Ni'].iloc[3])
    cu = (df['Cu'].iloc[3])

    return ni, cu, key
    
def custom_sort(s):
    num_part = int(s.split()[1].rstrip('М.XLSX'))
    if 'М' in s:
        if num_part == 1:
            return (0, -1)
        else:
            return (-num_part, -2)
            
    else:
        return (num_part, num_part)


def get_values(df):

    df['ПЛОТНОСТЬ'].iloc[0] = 3.11
    df['ПЛОТНОСТЬ'].iloc[1] = 1.85
    df['ПЛОТНОСТЬ'].iloc[2] = 2.95
    df['ПЛОТНОСТЬ'].iloc[3] = 3.54
    df['ПЛОТНОСТЬ'].iloc[6] = 2.75

    df = df[~((df[date] == 'ОР') | (df[date] == 'ПРОХ'))]


    M_row = df[df[date] == 'М']
    V_row = df[df[date] == 'В']
    I_row = df[df[date] == 'И']
    R_row = df[df[date] == 'Р']
    B_row = df[df[date] == 'ЗАКЛ']

    #1 col
    M_v = M_row['ОБЪЕМ'].values[0]                                  # B7
    V_v = V_row['ОБЪЕМ'].values[0]                                  # B9
    I_v = I_row['ОБЪЕМ'].values[0]                                  # B10
    R_v = R_row['ОБЪЕМ'].values[0]                                  # B11
    B_v = B_row['ОБЪЕМ'].values[0]                                  # B12
    Tovar_v = M_v + V_v + I_v + R_v + B_v                           # B5

    #2 col
    M_t = M_v * 3.54                                                # C7
    V_t = V_v * 3.11                                                # C9
    I_t = I_v * 2.95                                                # C10
    R_t = R_v * 2.75                                                # C11
    B_t = B_v * 1.85                                                # C12
    Tovar_t = M_t + V_t + I_t + R_t + B_t                           # C5

    R = (Tovar_t - (M_t + V_t))/Tovar_t*100

    #3 col
    M_Ni_value = M_row['Ni'].values[0]                              # D6
    M_Ni_weight = M_row['М_Ni'].values[0]                           # D7
    V_Ni_value = V_row['Ni'].values[0]                              # D8
    V_Ni_weight = V_row['М_Ni'].values[0]                           # D9
    Tovar_Ni_weight = M_t * M_Ni_value/100 + V_t * V_Ni_value/100   # D5
    Tovar_Ni_value = Tovar_Ni_weight / Tovar_t * 100                # D4

    #4 col
    M_Cu_value = M_row['Cu'].values[0]                              # E6
    M_Cu_weight = M_row['М_Cu'].values[0]                           # E7
    V_Cu_value = V_row['Cu'].values[0]                              # E8
    V_Cu_weight = V_row['М_Cu'].values[0]                           # E9
    Tovar_Cu_weight = M_t * M_Cu_value/100 + V_t * V_Cu_value/100   # E5
    Tovar_Cu_value = Tovar_Cu_weight / Tovar_t * 100                # E4

    #5 col
    M_Co_value = M_row['Co'].values[0]                              # F6
    M_Co_weight = M_row['М_Co'].values[0]                           # F7
    V_Co_value = V_row['Co'].values[0]                              # F8
    V_Co_weight = V_row['М_Co'].values[0]                           # F9
    Tovar_Co_weight = M_t * M_Co_value/100 + V_t * V_Co_value/100   # F5
    Tovar_Co_value = Tovar_Co_weight / Tovar_t * 100                # F4

    #7 col
    M_Pt_value = M_row['Pt'].values[0]                              # H6
    M_Pt_weight = M_row['М_Pt'].values[0]                           # H7
    V_Pt_value = V_row['Pt'].values[0]                              # H8
    V_Pt_weight = V_row['М_Pt'].values[0]                           # H9
    Tovar_Pt_weight = M_t * M_Pt_value/1000 + V_t * V_Pt_value/1000 # H5
    Tovar_Pt_value = Tovar_Pt_weight / Tovar_t * 1000               # H4

    #8 col
    M_Pd_value = M_row['Pd'].values[0]                              # I6
    M_Pd_weight = M_row['М_Pd'].values[0]                           # I7
    V_Pd_value = V_row['Pd'].values[0]                              # I8
    V_Pd_weight = V_row['М_Pd'].values[0]                           # I9
    Tovar_Pd_weight = M_t * M_Pd_value/1000 + V_t * V_Pd_value/1000 # I5
    Tovar_Pd_value = Tovar_Pd_weight / Tovar_t * 1000               # I4

    #9 col
    M_Rh_value = M_row['Rh'].values[0]                              # J6
    M_Rh_weight = M_row['М_Rh'].values[0]                           # J7
    V_Rh_value = V_row['Rh'].values[0]                              # J8
    V_Rh_weight = V_row['М_Rh'].values[0]                           # J9
    Tovar_Rh_weight = M_t * M_Rh_value/1000 + V_t * V_Rh_value/1000 # J5
    Tovar_Rh_value = Tovar_Rh_weight / Tovar_t * 1000               # J4

    #10 col
    M_Os_value = M_row['Os'].values[0]                              # K6
    M_Os_weight = M_row['М_Os'].values[0]                           # K7
    V_Os_value = V_row['Os'].values[0]                              # K8
    V_Os_weight = V_row['М_Os'].values[0]                           # K9
    Tovar_Os_weight = M_t * M_Os_value/1000 + V_t * V_Os_value/1000 # K5
    Tovar_Os_value = Tovar_Os_weight / Tovar_t * 1000               # K4

    #11 col
    M_Ir_value = M_row['Ir'].values[0]                              # L6
    M_Ir_weight = M_row['М_Ir'].values[0]                           # L7
    V_Ir_value = V_row['Ir'].values[0]                              # L8
    V_Ir_weight = V_row['М_Ir'].values[0]                           # L9
    Tovar_Ir_weight = M_t * M_Ir_value/1000 + V_t * V_Ir_value/1000 # L5
    Tovar_Ir_value = Tovar_Ir_weight / Tovar_t * 1000               # L4

    #12 col
    M_Ru_value = M_row['Ru'].values[0]                              # M6
    M_Ru_weight = M_row['М_Ru'].values[0]                           # M7
    V_Ru_value = V_row['Ru'].values[0]                              # M8
    V_Ru_weight = V_row['М_Ru'].values[0]                           # M9
    Tovar_Ru_weight = M_t * M_Ru_value/1000 + V_t * V_Ru_value/1000 # M5
    Tovar_Ru_value = Tovar_Ru_weight / Tovar_t * 1000               # M4

    #13 col
    M_Au_value = M_row['Au'].values[0]                              # N6
    M_Au_weight = M_row['М_Au'].values[0]                           # N7
    V_Au_value = V_row['Au'].values[0]                              # N8
    V_Au_weight = V_row['М_Au'].values[0]                           # N9
    Tovar_Au_weight = M_t * M_Au_value/1000 + V_t * V_Au_value/1000 # N5
    Tovar_Au_value = Tovar_Au_weight / Tovar_t * 1000               # N4

    #14 col
    M_Ag_value = M_row['Ag'].values[0]                              # O6
    M_Ag_weight = M_row['М_Ag'].values[0]                           # O7
    V_Ag_value = V_row['Ag'].values[0]                              # O8
    V_Ag_weight = V_row['М_Ag'].values[0]                           # O9
    Tovar_Ag_weight = M_t * M_Ag_value/1000 + V_t * V_Ag_value/1000 # O5
    Tovar_Ag_value = Tovar_Ag_weight / Tovar_t * 1000               # O4

    #15 col
    M_Se_value = M_row['Se'].values[0]                              # P6
    M_Se_weight = M_row['М_Se'].values[0]                           # P7
    V_Se_value = V_row['Se'].values[0]                              # P8
    V_Se_weight = V_row['М_Se'].values[0]                           # P9
    Tovar_Se_weight = M_t * M_Se_value/100 + V_t * V_Se_value/100   # P5
    Tovar_Se_value = Tovar_Se_weight / Tovar_t * 100                # P4

    #16 col
    M_Te_value = M_row['Te'].values[0]                              # Q6
    M_Te_weight = M_row['М_Te'].values[0]                           # Q7
    V_Te_value = V_row['Te'].values[0]                              # Q8
    V_Te_weight = V_row['М_Te'].values[0]                           # Q9
    Tovar_Te_weight = M_t * M_Te_value/100 + V_t * V_Te_value/100   # Q5
    Tovar_Te_value = Tovar_Te_weight / Tovar_t * 100                # Q4

    #17 col
    M_S_value = M_row['S'].values[0]                                # R6
    M_S_weight = M_row['М_S'].values[0]                             # R7
    V_S_value = V_row['S'].values[0]                                # R8
    V_S_weight = V_row['М_S'].values[0]                             # R9
    Tovar_S_weight = M_t * M_S_value/100 + V_t * V_S_value/100      # R5
    Tovar_S_value = Tovar_S_weight / Tovar_t * 100                  # R4

    MPG_weight = Tovar_Pt_weight + Tovar_Pd_weight + Tovar_Rh_weight + \
                 Tovar_Os_weight + Tovar_Ir_weight

    data = {
        'A1':'',             'A2':'',        'A3':'',             'A4':'ТОВАР',         'A5':'',         'A6':'М',         'A7':'',        'A8':'В',       'A9':'И', 'A10':'Р', 'A11':'ЗАКЛ', 'A12': '', 'A13':'Разубоживание', 'A14':'',
        'B1':'Объем,м3',     'B2':'',        'B3':'',             'B4':Tovar_v,         'B5':'',         'B6':M_v,         'B7':'',        'B8':V_v,       'B9':I_v, 'B10':R_v, 'B11':B_v,    'B12': '', 'B13':R,
        'C1':'Руда,т',       'C2':'',        'C3':'',             'C4':Tovar_t,         'C5':'',         'C6':M_t,         'C7':'',        'C8':V_t,       'C9':I_t, 'C10':R_t, 'C11':B_t,
        'D1':'Ni,т',         'D2':'%',       'D3':Tovar_Ni_value, 'D4':Tovar_Ni_weight, 'D5':M_Ni_value, 'D6':M_Ni_weight, 'D7':V_Ni_value,'D8':V_Ni_weight,
        'E1':'Cu,т',         'E2':'%',       'E3':Tovar_Cu_value, 'E4':Tovar_Cu_weight, 'E5':M_Cu_value, 'E6':M_Cu_weight, 'E7':V_Cu_value,'E8':V_Cu_weight,
        'F1':'Co,т',         'F2':'%',       'F3':Tovar_Co_value, 'F4':Tovar_Co_weight, 'F5':M_Co_value, 'F6':M_Co_weight, 'F7':V_Co_value,'F8':V_Co_weight,
        'G1':'Сумма МПГ,кг', 'G2':'',        'G3':'',             'G4':MPG_weight,      'G5':'',         'G6':'',          'G7':'',        'G8':'',        
        'H1':'Pt,кг',        'H2':'г/т',     'H3':Tovar_Pt_value, 'H4':Tovar_Pt_weight, 'H5':M_Pt_value, 'H6':M_Pt_weight, 'H7':V_Pt_value,'H8':V_Pt_weight,
        'I1':'Pd,кг',        'I2':'г/т',     'I3':Tovar_Pd_value, 'I4':Tovar_Pd_weight, 'I5':M_Pd_value, 'I6':M_Pd_weight, 'I7':V_Pd_value,'I8':V_Pd_weight,
        'J1':'Rh,кг',        'J2':'г/т',     'J3':Tovar_Rh_value, 'J4':Tovar_Rh_weight, 'J5':M_Rh_value, 'J6':M_Rh_weight, 'J7':V_Rh_value,'J8':V_Rh_weight,
        'K1':'Os,кг',        'K2':'г/т',     'K3':Tovar_Os_value, 'K4':Tovar_Os_weight, 'K5':M_Os_value, 'K6':M_Os_weight, 'K7':V_Os_value,'K8':V_Os_weight,
        'L1':'Ir,кг',        'L2':'г/т',     'L3':Tovar_Ir_value, 'L4':Tovar_Ir_weight, 'L5':M_Ir_value, 'L6':M_Ir_weight, 'L7':V_Ir_value,'L8':V_Ir_weight,
        'M1':'Ru,кг',        'M2':'г/т',     'M3':Tovar_Ru_value, 'M4':Tovar_Ru_weight, 'M5':M_Ru_value, 'M6':M_Ru_weight, 'M7':V_Ru_value,'M8':V_Ru_weight,
        'N1':'Au,кг',        'N2':'г/т',     'N3':Tovar_Au_value, 'N4':Tovar_Au_weight, 'N5':M_Au_value, 'N6':M_Au_weight, 'N7':V_Au_value,'N8':V_Au_weight,
        'O1':'Ag,кг',        'O2':'г/т',     'O3':Tovar_Ag_value, 'O4':Tovar_Ag_weight, 'O5':M_Ag_value, 'O6':M_Ag_weight, 'O7':V_Ag_value,'O8':V_Ag_weight,
        'P1':'Se,т',         'P2':'%',       'P3':Tovar_Se_value, 'P4':Tovar_Se_weight, 'P5':M_Se_value, 'P6':M_Se_weight, 'P7':V_Se_value,'P8':V_Se_weight,
        'Q1':'Te,т',         'Q2':'%',       'Q3':Tovar_Te_value, 'Q4':Tovar_Te_weight, 'Q5':M_Te_value, 'Q6':M_Te_weight, 'Q7':V_Te_value,'Q8':V_Te_weight,
        'R1':'S,т',          'R2':'%',       'R3':Tovar_S_value,  'R4':Tovar_S_weight,  'R5':M_S_value,  'R6':M_S_weight,  'R7':V_S_value, 'R8':V_S_weight}
        
    return data


def restyle_report(df):

    table_data = get_values(df)

    
    wb = Workbook()
    ws = wb.active

    for key, value in table_data.items():
        ws[key] = value


    header2 = NamedStyle(name='header2')
    header2.font = Font(bold=True)
    header2.border = Border(bottom=Side(border_style='thin'))
    header2.alignment = Alignment(horizontal='center', vertical='center')
    header1 = NamedStyle(name='header1')
    header1.font = Font(bold=True)
    header1.alignment = Alignment(horizontal='center', vertical='center')

    header_row1 = ws[1]
    for cell in header_row1:
        cell.style = header1
        
    header_row2 = ws[2]
    for cell in header_row2:
        cell.style = header2

    R_filler = PatternFill(patternType='solid', fgColor='FFC0CB')
    R_font = Font(color='00FF0000', size=11)
    ws['A13'].fill = R_filler
    ws['B13'].fill = R_filler
    ws['B13'].font = R_font



    columns = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','R']
    for column in columns:
        ws.column_dimensions[column].width = 10
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['G'].width = 15

    odd = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1',
            'A3','B3','C3','D3','E3','F3','G3','H3','I3','J3','K3','L3','M3','N3','O3','P3','Q3','R3',
            'A5','B5','C5','D5','E5','F5','G5','H5','I5','J5','K5','L5','M5','N5','O5','P5','Q5','R5',
            'A7','B7','C7','D7','E7','F7','G7','H7','I7','J7','K7','L7','M7','N7','O7','P7','Q7','R7',
            'A9','B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9',]


    color_line_1 = PatternFill(patternType='solid', fgColor='DCDCDC')
    for column in odd:
        ws[column].fill = color_line_1


    return wb, ws


def restyle_reports(tables_folder):

    listdir = os.listdir(tables_folder)

    if len(listdir) > 1:
        sorted_lst = sorted(listdir, key=custom_sort)
    else:
        sorted_lst = listdir

    for table_name in sorted_lst:

        table_path = os.path.join(tables_folder, table_name)
        df = pd.read_excel(table_path)
        wb, ws = restyle_report(df)
        wb.save(table_path)




def make_stats(tables_folder, planes):


    Ni_stamps = {}
    Cu_stamps = {}

    listdir = os.listdir(tables_folder)

    if len(listdir) > 1:
        sorted_lst = sorted(listdir, key=custom_sort)
    else:
        sorted_lst = listdir
  
    for table_name in sorted_lst:
        table_path = os.path.join(tables_folder, table_name)

        df = pd.read_excel(table_path)

        if len(listdir) > 1:
            ni, cu, key = metal_contents(table_name, df, planes)
            Ni_stamps[key] = ni
            Cu_stamps[key] = cu

    for table_name in sorted_lst:
        table_path = os.path.join(tables_folder, table_name)
        df = pd.read_excel(table_path)
        wb, ws = restyle_report(df)
        # Раздел: Весовые соотношения
        ws.merge_cells('A14:R15')
        ws['A14'] = 'Весовые соотношения'
        ws['A14'].alignment = Alignment(horizontal='center')
        filler = PatternFill(patternType='solid', fgColor='bdc3c7')
        for row in ws.iter_rows(min_row=14, max_row=15, min_col=1, max_col=18):
            for cell in row:
                cell.fill = filler
        #############################################
        table_data = get_values(df)
        ##############################################
        data = [['Элемент', 'Вес'],
               ['Ni', ws['D4'].value],
               ['Cu', ws['E4'].value],
               ['Co', ws['F4'].value],
               ['Pt', ws['H4'].value],
               ['Pd', ws['I4'].value],
               ['Rh', ws['J4'].value],
               ['Os', ws['K4'].value],
               ['Ir', ws['L4'].value],
               ['Ru', ws['M4'].value],
               ['Au', ws['N4'].value],
               ['Ag', ws['O4'].value],
               ['Te', ws['Q4'].value],
               ['S', ws['R4'].value/10]]
        ws.append(['',''])
        ws.append(['',''])
        for row in data:
            ws.append(row)
        chart_tovar = BarChart()
        chart_tovar.width = 10.1
        chart_tovar.height = 6.8
        values = Reference(ws, min_col=3, max_col=3, min_row=3, max_row=12)
        categories = Reference(ws, min_col=1, max_col=1, min_row=4, max_row=12)
        chart_tovar.add_data(values, titles_from_data=True)
        chart_tovar.set_categories(categories)
        chart_tovar.title = 'по типам'
        chart_bal = BarChart()
        chart_bal.width = 10.1
        chart_bal.height = 6.3
        values = Reference(ws, min_col=3, max_col=3, min_row=3, max_row=6)
        categories = Reference(ws, min_col=1, max_col=1, min_row=4, max_row=6)
        chart_bal.add_data(values, titles_from_data=True)
        chart_bal.set_categories(categories)
        chart_bal.title = 'товар/баланс'
        empty_plot = BarChart()
        empty_plot.width = 10.95
        empty_plot.height = 6.815
        ppie = ProjectedPieChart()
        ppie.height = 13.17
        ppie.width = 13.8
        ppie.type = 'pie'
        ppie.splitType = 'val'
        ppie.title = 'по элементам'
        labels = Reference(ws, min_col=1, max_col=1, min_row=18, max_row=30)
        data = Reference(ws, min_col=2, max_col=2, min_row=17, max_row=30)
        ppie.add_data(data, titles_from_data=True)
        ppie.set_categories(labels)
        data = [['Элемент', 'Вес'],
               ['Ni', ws['D4'].value],
               ['Cu', ws['E4'].value],
               ['Co', ws['F4'].value],
               ['МПГ', ws['G4'].value],
               ['Ru', ws['M4'].value],
               ['Au', ws['N4'].value],
               ['Ag', ws['O4'].value],
               ['Te', ws['Q4'].value],
               ['S', ws['R4'].value/10]]
        for row in data:
            ws.append(row)
        pie = PieChart()
        pie.type = 'pie'
        pie.splitType = 'val'
        pie.width = 10.95
        pie.height = 6.28
        pie.title = 'с объединенными МПГ'
        labels = Reference(ws, min_col=1, max_col=1, min_row=32, max_row=40)
        data = Reference(ws, min_col=2, max_col=2, min_row=31, max_row=40)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(chart_tovar, 'A16')
        ws.add_chart(chart_bal, 'A29')
        ws.add_chart(ppie, 'F16')
        ws.add_chart(pie, 'M16')
        ws.add_chart(empty_plot, 'M28')
        filler = PatternFill(patternType='solid', fgColor='5a5a5a')
        for row in ws.iter_rows(min_row=15, max_row=40, min_col=1, max_col=18):
            for cell in row:
                cell.fill = filler
        if len(listdir) > 1:
            chart = BarChart()
            chart.title= 'Изменчивость Ni/Cu по разрезам'
            chart.width = 35
            ws.cell(row=16, column=14, value='Ni')
            ws.cell(row=16, column=16, value='Cu')
            row = 17
            for key, value in Ni_stamps.items():
                ws.cell(row=row, column=13, value=key)
                ws.cell(row=row, column=14, value=value)
                row += 1
            data_1 = Reference(ws, min_col=14, max_col=14, min_row=16, max_row=row-1)
            categories_1 = Reference(ws, min_col=13, max_col=13, min_row=17, max_row=row-1)
            chart.add_data(data_1, titles_from_data=True)
            chart.set_categories(categories_1)
            s = chart.series[0]
            s.graphicalProperties.solidFill = 'FBCEB1'
            row = 17
            for key, value in Cu_stamps.items():
                ws.cell(row=row, column=15, value=key)
                ws.cell(row=row, column=16, value=value)
                row += 1
            data_2 = Reference(ws, min_col=16, max_col=16, min_row=16, max_row=row-1)
            categories_2 = Reference(ws, min_col=15, max_col=15, min_row=17, max_row=row-1)
            chart.add_data(data_2, titles_from_data=True)
            chart.set_categories(categories_2)
            s = chart.series[1]
            s.graphicalProperties.solidFill = '00FF00'
            ws.add_chart(chart, "A41")
        wb.save(table_path)

        
        
